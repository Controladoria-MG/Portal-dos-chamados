"""
gerar_relatorio_coordenadores.py
─────────────────────────────────
Cruza a planilha de colaboradores ativos com "data/Att Base/Coordenadores.xlsx"
(por Nome Completo) e gera um novo arquivo com duas abas:

  - "Colaboradores x Coordenadores": todos os colaboradores, com o
    coordenador correspondente (em branco quando não encontrado).
  - "Sem Coordenador": só os colaboradores sem coordenador identificado,
    para servir de lista de pendências.

O relatório de colaboradores e o Coordenadores.xlsx não são alterados.

Rodar manualmente, sob demanda, sempre que um dos dois for atualizado:
    python backend/gerar_relatorio_coordenadores.py
"""

import re
from pathlib import Path

import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
ATT_DIR  = BASE_DIR / "data" / "Att Base"
OUT_DIR  = BASE_DIR / "data" / "coordenadores"
OUTPUT   = OUT_DIR / "Colaboradores_Coordenadores.xlsx"

HEADERS = ["Nome Completo", "Nome de Exibição", "Departamento", "Coordenador"]
COL_WIDTHS = {1: 34, 2: 22, 3: 30, 4: 26}


def _find_file(pattern):
    matches = list(ATT_DIR.glob(pattern))
    if not matches:
        raise FileNotFoundError(f"Nenhum arquivo '{pattern}' encontrado em {ATT_DIR}")
    return max(matches, key=lambda p: p.stat().st_mtime)


def _norm(nome):
    if not nome:
        return ""
    return re.sub(r"\s+", " ", str(nome).strip().upper())


def _montar_aba(ws, headers, linhas, nome_tabela):
    ws.append(headers)
    for linha in linhas:
        ws.append(list(linha))

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = COL_WIDTHS.get(idx, 20)

    if ws.max_row > 1:
        tbl = Table(displayName=nome_tabela, ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}")
        tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(tbl)


def main():
    colab_path = _find_file("*olaboradores*.xlsx")
    coord_path = _find_file("*oordenadores*.xlsx")

    print(f"Colaboradores: {colab_path.name}")
    print(f"Coordenadores: {coord_path.name}")

    wb_coord = openpyxl.load_workbook(coord_path, data_only=True)
    ws_coord = wb_coord.active
    coord_map = {}
    for row in ws_coord.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        coord_map[_norm(row[0])] = row[3]  # coluna "Coordenador"
    wb_coord.close()

    wb_colab = openpyxl.load_workbook(colab_path, data_only=True)
    ws_colab = wb_colab.active

    linhas = []
    for row in ws_colab.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nome_completo = row[0]
        nome_exib     = row[1]
        depto         = row[5]
        coordenador   = coord_map.get(_norm(nome_completo), "")
        linhas.append((nome_completo, nome_exib, depto, coordenador))
    wb_colab.close()

    sem_coord = [l[:3] for l in linhas if not l[3]]

    wb_out = openpyxl.Workbook()
    ws1 = wb_out.active
    ws1.title = "Colaboradores x Coordenadores"
    _montar_aba(ws1, HEADERS, linhas, "TabelaColaboradoresCoordenadores")

    ws2 = wb_out.create_sheet("Sem Coordenador")
    _montar_aba(ws2, HEADERS[:3], sem_coord, "TabelaSemCoordenador")

    OUT_DIR.mkdir(exist_ok=True)
    wb_out.save(OUTPUT)

    print()
    print(f"{len(linhas)} colaboradores processados")
    print(f"{len(linhas) - len(sem_coord)} com coordenador identificado")
    print(f"{len(sem_coord)} sem coordenador -> aba 'Sem Coordenador'")
    print(f"Arquivo gerado: {OUTPUT}")


if __name__ == "__main__":
    main()
