# ============================================================
# MG Contécnica – Extrator + Atualizador de Base de Chamados
# Versão 1.2 (unificado)
# - Download automático via Selenium
# - Seleção múltipla de departamentos (select2)
# - Opção sequencial (todas as extrações de uma vez)
# - Atualização automática do base.xlsx ao final
# ============================================================

import os
import re
import sys
import time
import json
import subprocess
import argparse
from pathlib import Path
from datetime import datetime

# ============================================================
# CONSOLE – fixar tamanho e codepage ANTES de qualquer print,
# para o rich não calcular a largura errada e desalinhar a
# tabela do menu inicial.
# ============================================================

if sys.platform == "win32":
    import ctypes
    os.system("mode con: cols=130 lines=45")
    ctypes.windll.kernel32.SetConsoleOutputCP(65001)
    ctypes.windll.kernel32.SetConsoleCP(65001)
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ============================================================
# GARANTIR DEPENDÊNCIAS
# ============================================================

def _instalar(pacote):
    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", pacote, "--quiet"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text
    from rich import box
    from rich.prompt import Prompt
    from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
    from rich.align import Align
except ImportError:
    print("Instalando dependência 'rich'...")
    _instalar("rich")
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table
    from rich.text import Text
    from rich import box
    from rich.prompt import Prompt
    from rich.progress import Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
    from rich.align import Align

try:
    import openpyxl
    from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando dependência 'openpyxl'...")
    _instalar("openpyxl")
    import openpyxl
    from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

console = Console(legacy_windows=False)

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# ============================================================
# CONFIGURAÇÕES – EXTRAÇÃO
# ============================================================

USUARIO  = "rjuan"
SENHA    = "Palmeiras!"

PASTA_DOWNLOAD    = r"C:\Users\gamaral\Desktop\Python\Chamados\data\Att Base"
ARQUIVO_PROTEGIDO = "Relatório Colaboradores Ativos_07_07_2026, 17_41_11.xlsx"
REPO_DIR          = r"C:\Users\gamaral\Desktop\Python\Chamados"

TIMEOUT    = 40
URL_LOGIN  = "https://aplicativo.mgcontecnica.com.br/#/login"
DATA_INICIO = "01/01/2026"

TIPOS_EXTRACAO = {
    "1": {
        "label":        "Geral",
        "descricao":    "Todos os departamentos operacionais",
        "departamentos": [
            "DP - DEPTO. PESSOAL",
            "EF - VAREJO",
            "EF - INDUSTRIA",
            "CTB - CONTÁBIL VAREJO",
            "CTB - CONTÁBIL INDUSTRIA",
            "CTB - CONTÁBIL HOLDING",
        ],
    },
    "2": {
        "label":        "Controladoria",
        "descricao":    "Apenas Controladoria",
        "departamentos": ["CONTROLADORIA"],
    },
    "3": {
        "label":        "GC",
        "descricao":    "Gerência de Contas + GC Administrativo",
        "departamentos": ["GERENCIA DE CONTAS", "GC - ADMINISTRATIVO"],
    },
    "4": {
        "label":        "Paralegal",
        "descricao":    "Apenas Paralegal",
        "departamentos": ["PARALEGAL"],
    },
}

# ============================================================
# CONFIGURAÇÕES – BASE
# ============================================================

_ATT_DIR = Path(PASTA_DOWNLOAD)
_DATA_DIR = _ATT_DIR.parent
OUTPUT_BASE = _DATA_DIR / "base.xlsx"

DEPTS_RETORNADOS = {"gerencia de contas", "gc - administrativo"}
STATUS_ENTREGUE  = "entregue ao solicitante"
STATUS_ENCERRADO = "encerrado"
STATUS_DEVOLVIDO_SOLICITANTE = "devolvido para solicitante"

COL_MAP = [
    (0,  0),   # A → A  | Id
    (1,  1),   # B → B  | Categoria
    (2,  2),   # C → C  | Assunto
    (3,  3),   # D → D  | Departamento Solicitante
    (4,  4),   # E → E  | Solicitação
    (5,  5),   # F → F  | IdCliente
    (6,  6),   # G → G  | Cliente
    (8,  7),   # I → H  | Responsável (com lookup)
    (9,  8),   # J → I  | Departamento Responsavel
    (13, 9),   # N → J  | Data Cadastro
    (14, 10),  # O → K  | Prazo Vencimento
    (15, 11),  # P → L  | Status
    (16, 12),  # Q → M  | Solicitante (com lookup)
    (17, 13),  # R → N  | Inicio Atend.
    (18, 14),  # S → O  | DataPrevisaoAtendimento
    (19, 15),  # T → P  | Data Entrega
    (20, 16),  # U → Q  | Responsável pela conclusão
    (21, 17),  # V → R  | Ultimo Comentário
]

LOOKUP_COLS = {8, 16}

HEADERS = [
    "Id", "Categoria", "Assunto", "Departamento Solicitante",
    "Solicitação", "IdCliente", "Cliente", "Responsável",
    "Departamento Responsavel", "Data Cadastro", "Prazo Vencimento",
    "Status", "Solicitante", "Inicio Atend.", "DataPrevisaoAtendimento",
    "Data Entrega", "Responsável pela conclusão", "Ultimo Comentário",
    "Retornado", "Departamento Responsavel Original", "Coordenador",
]

COL_WIDTHS = {
    'A': 10,    'B': 11.86, 'C': 10.43, 'D': 26.29,
    'E': 13.0,  'F': 11.57, 'G': 9.86,  'H': 14.71,
    'I': 28.14, 'J': 15.86, 'K': 19.29, 'L': 12.86,
    'M': 12.86, 'N': 14.43, 'O': 14.43, 'P': 15.57,
    'Q': 28.57, 'R': 20.29, 'S': 12.0,  'T': 28.14,
    'U': 22.0,
}

# ============================================================
# MENU – TIPO DE EXTRAÇÃO
# ============================================================

def perguntar_extracao():
    console.print()
    console.rule("[bold cyan]Tipo de Extração[/bold cyan]")
    console.print()

    table = Table(box=box.ROUNDED, border_style="dim", show_header=False, padding=(0, 2))
    table.add_column("Key",     style="bold cyan",  width=5)
    table.add_column("Tipo",    style="bold white",  width=18)
    table.add_column("Detalhe", style="dim")

    for k, v in TIPOS_EXTRACAO.items():
        deptos = ", ".join(v["departamentos"])
        table.add_row(f"[{k}]", v["label"],
                      v["descricao"] + f"\n[dim italic]{deptos}[/dim italic]")

    table.add_row("[5]", "Todas (seq.)",
                  f"Executa as {len(TIPOS_EXTRACAO)} extrações em sequência — gera {len(TIPOS_EXTRACAO)} arquivos separados")

    console.print(Align.center(table))
    console.print()

    while True:
        opcao = Prompt.ask("[bold cyan]  Digite a opção[/bold cyan]").strip()
        if opcao in TIPOS_EXTRACAO:
            escolha = TIPOS_EXTRACAO[opcao]
            console.print("  [green]✔ Extração:[/green] [bold]" + escolha["label"] + "[/bold]")
            return opcao, escolha["label"], escolha["departamentos"]
        if opcao == "5":
            console.print(f"  [green]✔ Extração:[/green] [bold]Todas (sequencial — {len(TIPOS_EXTRACAO)} arquivos)[/bold]")
            return "5", "Todas", None
        console.print("  [red]Opção inválida. Tente novamente.[/red]")

# ============================================================
# MENU – MODO DE VISUALIZAÇÃO
# ============================================================

def perguntar_modo():
    console.print()
    console.rule("[bold cyan]Modo de Execução[/bold cyan]")
    console.print()

    table = Table(box=box.ROUNDED, border_style="dim", show_header=False, padding=(0, 2))
    table.add_column("Key",  style="bold cyan", width=5)
    table.add_column("Modo", style="bold white")
    table.add_column("Desc", style="dim")

    table.add_row("[1]", "Com janela", "Abre o Chrome visível — você acompanha tudo em tempo real")
    table.add_row("[2]", "Sem janela", "Roda em segundo plano (headless) — mais rápido")

    console.print(Align.center(table))
    console.print()

    while True:
        opcao = Prompt.ask("[bold cyan]  Digite a opção[/bold cyan]").strip()
        if opcao == "1":
            console.print("  [green]✔ Modo:[/green] [bold]Com janela (visível)[/bold]")
            return False
        if opcao == "2":
            console.print("  [green]✔ Modo:[/green] [bold]Sem janela (headless)[/bold]")
            return True
        console.print("  [red]Opção inválida. Digite 1 ou 2.[/red]")

# ============================================================
# UTILITÁRIOS
# ============================================================

def hoje_str():
    return datetime.now().strftime("%d/%m/%Y")

def agora_str():
    return datetime.now().strftime("%d-%m-%y %H%M")

# ============================================================
# PASTA / ARQUIVOS
# ============================================================

def limpar_pasta_download(pasta, progress, task):
    progress.update(task, description="[cyan]Removendo arquivos de chamados anteriores...")
    removidos = 0
    try:
        for arq in os.listdir(pasta):
            if (arq.lower().endswith(".xlsx")
                    and arq.startswith("Chamados -")
                    and arq != ARQUIVO_PROTEGIDO):
                try:
                    os.remove(os.path.join(pasta, arq))
                    removidos += 1
                except Exception:
                    pass
    except FileNotFoundError:
        pass
    progress.update(task, description=f"[cyan]{removidos} arquivo(s) antigo(s) removido(s).")
    time.sleep(0.5)


def snapshot_xlsx(pasta: str) -> set:
    try:
        return {
            os.path.join(pasta, arq)
            for arq in os.listdir(pasta)
            if arq.lower().endswith(".xlsx")
        }
    except FileNotFoundError:
        return set()


def aguardar_download(pasta, antes: set, progress, task, timeout=600):
    inicio         = time.time()
    ultimo_arquivo = None
    ultimo_tamanho = -1
    tempo_estavel  = 0

    while time.time() - inicio < timeout:
        arquivos_temp = [
            arq for arq in os.listdir(pasta)
            if arq.endswith(".crdownload") or arq.endswith(".tmp")
        ]
        if arquivos_temp:
            progress.update(task, description="[cyan]Arquivo temporário detectado, aguardando...")
            time.sleep(1)
            continue

        novos = [
            os.path.join(pasta, arq)
            for arq in os.listdir(pasta)
            if arq.lower().endswith(".xlsx")
            and os.path.join(pasta, arq) not in antes
        ]

        if not novos:
            progress.update(task, description="[cyan]Aguardando arquivo aparecer na pasta...")
            time.sleep(1)
            continue

        arquivo_recente = max(novos, key=os.path.getctime)

        try:
            tamanho_atual = os.path.getsize(arquivo_recente)

            if ultimo_arquivo == arquivo_recente and ultimo_tamanho == tamanho_atual:
                tempo_estavel += 1
                progress.update(task, description=f"[cyan]Verificando estabilidade ({tempo_estavel}/3)...")
            else:
                tempo_estavel = 0
                progress.update(task, description="[cyan]Arquivo crescendo, aguardando...")

            ultimo_arquivo = arquivo_recente
            ultimo_tamanho = tamanho_atual

            if tempo_estavel >= 3:
                return arquivo_recente

        except Exception:
            pass

        time.sleep(1)

    return None

# ============================================================
# DRIVER
# ============================================================

def iniciar_driver(pasta_download, headless=True):
    os.makedirs(pasta_download, exist_ok=True)

    options = Options()
    if headless:
        options.add_argument("--headless=new")
        # Chrome recente às vezes abre uma janela em branco mesmo em modo
        # headless — joga essa janela pra fora da tela em vez de deixá-la
        # aparecer no meio da execução automática.
        options.add_argument("--window-position=-32000,-32000")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")

    prefs = {
        "download.default_directory":                                 os.path.abspath(pasta_download),
        "download.prompt_for_download":                               False,
        "download.directory_upgrade":                                 True,
        "safebrowsing.enabled":                                       True,
        "safebrowsing.disable_download_protection":                   True,
        "profile.default_content_setting_values.automatic_downloads": 1,
    }
    options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())
    driver  = webdriver.Chrome(service=service, options=options)

    driver.execute_cdp_cmd(
        "Page.setDownloadBehavior",
        {"behavior": "allow", "downloadPath": os.path.abspath(pasta_download)},
    )
    return driver

# ============================================================
# AUTOMAÇÃO WEB
# ============================================================

def fazer_login(driver, progress, task):
    progress.update(task, description="[cyan]Abrindo página de login...")
    driver.get(URL_LOGIN)
    wait = WebDriverWait(driver, TIMEOUT)

    progress.update(task, description="[cyan]Preenchendo credenciais...")
    campo_usuario = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "input[formcontrolname='usuario'], input[id='usuario']")
        )
    )
    campo_usuario.clear()
    campo_usuario.send_keys(USUARIO)

    campo_senha = wait.until(
        EC.presence_of_element_located(
            (By.CSS_SELECTOR, "input#senha[formcontrolname='senha']")
        )
    )
    campo_senha.clear()
    campo_senha.send_keys(SENHA)

    driver.find_element(By.CSS_SELECTOR, "button[type='submit']").click()

    progress.update(task, description="[cyan]Aguardando autenticação...")
    wait.until(EC.url_contains("/home"))


def navegar_para_relatorio(driver, progress, task):
    wait = WebDriverWait(driver, TIMEOUT)

    progress.update(task, description="[cyan]Abrindo MG Controle...")
    mg_controle = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH, "//div[@title='http://intranetmg/Aplicativos/Geral/Controle/']")
        )
    )
    driver.execute_script("arguments[0].click();", mg_controle)
    time.sleep(4)
    if len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])

    progress.update(task, description="[cyan]Acessando Chamados...")
    btn_chamados = wait.until(
        EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_ChamadoGeralLinkButton"))
    )
    driver.execute_script("arguments[0].click();", btn_chamados)
    time.sleep(3)
    if len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])

    progress.update(task, description="[cyan]Acessando Relatórios...")
    btn_relatorios = wait.until(
        EC.element_to_be_clickable((By.ID, "ContentPlaceHolder1_RelatorioLinkButton"))
    )
    driver.execute_script("arguments[0].click();", btn_relatorios)
    time.sleep(3)
    if len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])

    progress.update(task, description="[cyan]Abrindo Relatório Chamados...")
    btn_rel_chamados = wait.until(
        EC.element_to_be_clickable((By.ID, "btnRelChamados"))
    )
    driver.execute_script("arguments[0].click();", btn_rel_chamados)
    time.sleep(4)
    if len(driver.window_handles) > 1:
        driver.switch_to.window(driver.window_handles[-1])


def limpar_departamentos(driver, qtd_selecionados, progress, task):
    wait = WebDriverWait(driver, TIMEOUT)

    progress.update(task, description="[cyan]Limpando seleções anteriores...")
    select2 = wait.until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, ".select2-container"))
    )
    driver.execute_script("arguments[0].click();", select2)
    time.sleep(1)

    campo_busca = wait.until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, ".select2-input"))
    )
    for _ in range(qtd_selecionados + 1):
        campo_busca.send_keys(Keys.BACKSPACE)
        time.sleep(0.3)

    campo_busca.send_keys(Keys.ESCAPE)
    time.sleep(0.5)


def selecionar_departamentos(driver, departamentos, progress, task):
    wait = WebDriverWait(driver, TIMEOUT)

    for depto in departamentos:
        progress.update(task, description=f"[cyan]Selecionando: {depto}...")
        select2 = wait.until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".select2-container"))
        )
        driver.execute_script("arguments[0].click();", select2)
        time.sleep(1)

        campo_busca = wait.until(
            EC.visibility_of_element_located((By.CSS_SELECTOR, ".select2-input"))
        )
        campo_busca.clear()
        campo_busca.send_keys(depto)
        time.sleep(2)
        campo_busca.send_keys(Keys.ENTER)
        time.sleep(1)

    progress.update(task, description=f"[cyan]{len(departamentos)} departamento(s) selecionado(s).")
    time.sleep(1)


def preencher_data_campo(driver, campo_id, data_str, progress, task, descricao):
    wait = WebDriverWait(driver, TIMEOUT)

    progress.update(task, description=f"[cyan]{descricao}...")
    campo = wait.until(EC.element_to_be_clickable((By.ID, campo_id)))

    driver.execute_script("arguments[0].click();", campo)
    time.sleep(1)
    campo.send_keys(Keys.ESCAPE)
    time.sleep(0.5)
    campo.send_keys(Keys.CONTROL + "a")
    campo.send_keys(Keys.DELETE)
    time.sleep(0.3)

    for char in data_str:
        campo.send_keys(char)
        time.sleep(0.05)

    campo.send_keys(Keys.TAB)
    time.sleep(0.5)

    valor = driver.execute_script("return arguments[0].value;", campo)
    if valor != data_str:
        driver.execute_script(f"arguments[0].value='{data_str}';", campo)
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('change', {bubbles:true}));", campo
        )
        time.sleep(0.3)


def marcar_encerrados(driver, progress, task):
    """Garante que a checkbox 'Encerrados' esteja marcada antes de gerar o
    relatório. Sem ela, os chamados já encerrados (que vêm numa aba extra
    'Encerrados' no arquivo baixado) não são incluídos na extração e ficam
    de fora da contagem de baixados do histórico mensal."""
    wait = WebDriverWait(driver, TIMEOUT)
    progress.update(task, description="[cyan]Marcando 'Encerrados'...")
    try:
        checkbox = wait.until(
            EC.presence_of_element_located(
                (By.XPATH,
                 "//label[contains(@class,'checkbox')]"
                 "[contains(normalize-space(.), 'Encerrados')]"
                 "//input[@type='checkbox']")
            )
        )
        if not checkbox.is_selected():
            driver.execute_script("arguments[0].click();", checkbox)
            time.sleep(0.3)
    except Exception:
        progress.update(task, description="[yellow]Checkbox 'Encerrados' não encontrada — seguindo sem marcar.")
        time.sleep(1)


def preencher_datas(driver, data_inicio, data_fim, progress, task):
    preencher_data_campo(driver, "txtAberturaInicio", data_inicio, progress, task,
                         f"Preenchendo data inicio: {data_inicio}")
    preencher_data_campo(driver, "txtAberturaFim",    data_fim,    progress, task,
                         f"Preenchendo data fim: {data_fim}")


def gerar_relatorio(driver, pasta_download, progress, task):
    wait  = WebDriverWait(driver, TIMEOUT)
    os.makedirs(pasta_download, exist_ok=True)
    antes = snapshot_xlsx(pasta_download)

    progress.update(task, description="[cyan]Clicando em Gerar Relatório...")
    btn_gerar = wait.until(
        EC.element_to_be_clickable(
            (By.XPATH,
             "//a[contains(@class,'btn') and contains(@class,'btnVerde') "
             "and contains(text(),'Gerar Relatório')]")
        )
    )
    driver.execute_script("arguments[0].click();", btn_gerar)
    return aguardar_download(pasta_download, antes, progress, task)


def renomear_arquivo(caminho_original, pasta_download, label="Geral"):
    timestamp    = agora_str()
    novo_nome    = f"Chamados - {label} - {timestamp}.xlsx"
    novo_caminho = os.path.join(pasta_download, novo_nome)
    os.rename(caminho_original, novo_caminho)
    return novo_caminho, novo_nome

# ============================================================
# PROCESSAMENTO BASE
# ============================================================

def _norm_nome(nome):
    if not nome:
        return ""
    return re.sub(r"\s+", " ", str(nome).strip().upper())


def _find_file(folder: Path, pattern: str, required=True):
    matches = list(folder.glob(pattern))
    if not matches:
        if required:
            raise FileNotFoundError(
                f"Nenhum arquivo '{pattern}' encontrado em {folder}"
            )
        return None
    return max(matches, key=lambda p: p.stat().st_mtime)


def _dept_group(nome):
    """Agrupa o Departamento Responsavel do jeito que o portal exibe —
    GC - Administrativo sempre unificado em GERENCIA DE CONTAS, e
    PL - Auditoria (nome do depto no relatório de origem) exibido como
    PARALEGAL. Ver deptGroupName() em script.js (mesma regra, front e back)."""
    dept = str(nome).strip().upper() if nome else ''
    if dept == 'GC - ADMINISTRATIVO':
        return 'GERENCIA DE CONTAS'
    if dept == 'PL - AUDITORIA':
        return 'PARALEGAL'
    return dept or 'SEM DEPARTAMENTO'


def _classificar_prazo(data_entrega, prazo):
    """Chamado fechado (Entregue/Encerrado): dentro ou fora do Prazo
    Vencimento. Sem data de entrega ou sem prazo registrado, considera
    dentro do prazo por padrão (não há como provar atraso)."""
    if not isinstance(data_entrega, datetime) or not isinstance(prazo, datetime):
        return 'atendido_dentro'
    return 'atendido_dentro' if data_entrega.date() <= prazo.date() else 'atendido_fora'


def _processar_relatorio(ws, label, colab_map, coordenador_map, ws_out, detalhe):
    linhas = ignoradas = retornados = 0

    def resolver(valor):
        if not valor:
            return valor
        return colab_map.get(str(valor).strip().lower(), str(valor).strip())

    nao_encontrados = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        status_val = str(row[15]).strip().lower() if row[15] else ''
        dept_sol   = str(row[3]).strip().lower()  if row[3]  else ''
        # "Encerrado" conta como baixado no histórico igual a "Entregue Ao
        # Solicitante", mas NUNCA vira "Retornado" para a GC — só chamados
        # entregues normalmente entram na aba de retornados. Encerrado,
        # mesmo sendo da GC, é sempre descartado da base (não aparece em
        # lugar nenhum do portal).
        eh_entregue = (status_val == STATUS_ENTREGUE)
        eh_encerrado = (status_val == STATUS_ENCERRADO)
        eh_fechado  = eh_entregue or eh_encerrado
        eh_gc       = (dept_sol in DEPTS_RETORNADOS)
        manter_como_retornado = eh_entregue and eh_gc

        # Detalhe mensal por departamento/categoria (para o relatório do
        # portal) — computado sobre TODAS as linhas do relatório bruto, antes
        # de qualquer descarte, para não perder os chamados finalizados fora
        # da GC que a base.xlsx não guarda. O mês do chamado é sempre o mês
        # de ABERTURA (Data Cadastro) — não o de atendimento — para que
        # "abrir os chamados daquele mês" sempre mostre o mesmo conjunto,
        # cada um com seu status final (atendido dentro/fora do prazo, ou
        # ainda pendente). Departamento = Departamento Responsavel (coluna J
        # do relatório de origem), agrupado como no portal (deptGroupName) —
        # e, em "Devolvido para Solicitante", reatribuído para o depto
        # solicitante (coluna D), igual ao que acontece com nova[8] mais
        # abaixo para a base.xlsx. Sem isso, um chamado devolvido contava no
        # relatório sob o depto original em vez de para quem ele voltou,
        # divergindo da contagem de pendentes mostrada na tela de Chamados.
        if status_val == STATUS_DEVOLVIDO_SOLICITANTE:
            depto_raw = row[3] if len(row) > 3 else None
        else:
            depto_raw = row[9] if len(row) > 9 else None
        depto = _dept_group(depto_raw)

        data_cad_raw = row[13] if len(row) > 13 else None
        if isinstance(data_cad_raw, datetime):
            mes_chave    = data_cad_raw.strftime('%Y-%m')
            categoria    = str(row[1]).strip() if row[1] else 'Sem Categoria'
            prazo_raw    = row[14] if len(row) > 14 else None
            data_ent_raw = row[19] if len(row) > 19 else None

            status_final = _classificar_prazo(data_ent_raw, prazo_raw) if eh_fechado else 'pendente'

            grupo = (detalhe.setdefault(depto, {})
                            .setdefault(mes_chave, {})
                            .setdefault(categoria, {
                                'atendidosDentro': 0, 'atendidosFora': 0, 'pendentes': 0,
                                'chamados': [],
                            }))
            if status_final == 'atendido_dentro':
                grupo['atendidosDentro'] += 1
            elif status_final == 'atendido_fora':
                grupo['atendidosFora'] += 1
            else:
                grupo['pendentes'] += 1

            grupo['chamados'].append({
                'id':           row[0],
                'cliente':      row[6] if len(row) > 6 else None,
                'dataCadastro': data_cad_raw.strftime('%Y-%m-%d'),
                'prazo':        prazo_raw.strftime('%Y-%m-%d') if isinstance(prazo_raw, datetime) else None,
                'dataEntrega':  data_ent_raw.strftime('%Y-%m-%d') if isinstance(data_ent_raw, datetime) else None,
                'status':       status_final,
            })

        if eh_fechado and not manter_como_retornado:
            ignoradas += 1
            continue

        nova = [None] * len(HEADERS)
        for col_src, col_dst in COL_MAP:
            valor = row[col_src] if col_src < len(row) else None
            if col_src in LOOKUP_COLS:
                resolvido = resolver(valor)
                # Só entra como "não encontrado" se parecer um usuário AD de
                # verdade (sem espaço) — o relatório de origem passou a
                # trazer o nome de exibição pronto em algumas linhas, e esse
                # nome nunca vai bater com a chave (usuário AD) do colab_map,
                # o que inflaria a lista de "não encontrados" à toa mesmo
                # com o valor certo já exibido (via fallback do resolver).
                bruto = str(valor).strip() if valor else ''
                if valor and resolvido == bruto and ' ' not in bruto:
                    nao_encontrados.add(bruto)
                valor = resolvido
            nova[col_dst] = valor

        # "Devolvido para Solicitante": a pendência passa a ser do depto
        # solicitante (coluna D), não mais do depto responsável (coluna I).
        # O depto responsável original é preservado na coluna de rastreabilidade.
        if status_val == STATUS_DEVOLVIDO_SOLICITANTE:
            nova[19] = nova[8]
            nova[8]  = nova[3]

        nova[18] = "SIM" if manter_como_retornado else "NÃO"
        if manter_como_retornado:
            retornados += 1

        # Coordenador do responsável pelo chamado (via Coordenadores.xlsx,
        # cruzado pelo Nome de Exibição já resolvido na coluna H).
        nova[20] = coordenador_map.get(_norm_nome(nova[7]), "")

        ws_out.append(nova)
        linhas += 1

    console.print(
        f"  [dim]{label}:[/dim] "
        f"[green]{linhas}[/green] incluídos  "
        f"[yellow]{ignoradas}[/yellow] ignorados  "
        f"[cyan]{retornados}[/cyan] retornados"
    )
    return linhas, ignoradas, retornados, nao_encontrados


def atualizar_base():
    console.print()
    console.rule("[bold cyan]Atualizando base.xlsx[/bold cyan]")
    console.print()

    att_dir = _ATT_DIR

    colab_path = _find_file(att_dir, "*olaboradores*.xlsx")
    coord_path = _find_file(att_dir, "*oordenadores*.xlsx",             required=False)
    gc_path    = _find_file(att_dir, "*Chamados - GC*.xlsx",            required=False)
    ctr_path   = _find_file(att_dir, "*Chamados - Controladoria*.xlsx", required=False)
    pl_path    = _find_file(att_dir, "*Chamados - Paralegal*.xlsx",     required=False)

    gc_name  = gc_path.name  if gc_path  else None
    ctr_name = ctr_path.name if ctr_path else None
    pl_name  = pl_path.name  if pl_path  else None

    chamados_paths = [
        p for p in att_dir.glob("*Chamados*.xlsx")
        if p.name != gc_name and p.name != ctr_name and p.name != pl_name
    ]
    if not chamados_paths:
        raise FileNotFoundError(
            f"Nenhum arquivo Chamados (Geral) encontrado em {att_dir}"
        )

    console.print(f"  [dim]Colaboradores:[/dim] {colab_path.name}")
    console.print(f"  [dim]Coordenadores:[/dim] {coord_path.name if coord_path else '[yellow]não encontrado[/yellow]'}")
    for p in sorted(chamados_paths):
        console.print(f"  [dim]Chamados:[/dim]      {p.name}")
    console.print(f"  [dim]GC:[/dim]            {gc_name  or '[yellow]não encontrado[/yellow]'}")
    console.print(f"  [dim]Controladoria:[/dim] {ctr_name or '[yellow]não encontrado[/yellow]'}")
    console.print(f"  [dim]Paralegal:[/dim]     {pl_name  or '[yellow]não encontrado[/yellow]'}")
    console.print()

    # Mapa de colaboradores
    wb_colab  = openpyxl.load_workbook(colab_path, read_only=True, data_only=True)
    colab_map = {}
    for row in wb_colab.active.iter_rows(min_row=2, values_only=True):
        nome_exib, usuario = row[1], row[2]
        if usuario and nome_exib:
            colab_map[str(usuario).strip().lower()] = str(nome_exib).strip()
    wb_colab.close()
    console.print(f"  [dim]{len(colab_map)} colaboradores carregados.[/dim]")

    # Mapa de coordenadores (Nome de Exibição → Coordenador). Opcional —
    # se o arquivo não existir, a coluna Coordenador fica em branco.
    coordenador_map = {}
    if coord_path:
        wb_coord = openpyxl.load_workbook(coord_path, read_only=True, data_only=True)
        for row in wb_coord.active.iter_rows(min_row=2, values_only=True):
            nome_exib   = row[1] if len(row) > 1 else None
            coordenador = row[3] if len(row) > 3 else None
            if nome_exib and coordenador:
                coordenador_map[_norm_nome(nome_exib)] = str(coordenador).strip()
        wb_coord.close()
    console.print(
        f"  [dim]{len(coordenador_map)} coordenadores carregados"
        f"{' (' + coord_path.name + ')' if coord_path else ' — Coordenadores.xlsx não encontrado'}.[/dim]"
    )
    console.print()

    # Workbook de saída
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Base"
    ws_out.append(HEADERS)

    total_linhas = total_ignoradas = total_retornados = 0
    todos_nao_encontrados: set = set()
    detalhe: dict = {}

    fontes = [(p.stem, p) for p in sorted(chamados_paths)]
    if gc_path:  fontes.append(("GC",            gc_path))
    if ctr_path: fontes.append(("Controladoria", ctr_path))
    if pl_path:  fontes.append(("Paralegal",     pl_path))

    for label, path in fontes:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lin, ign, ret, nao = _processar_relatorio(
            wb.active, label, colab_map, coordenador_map, ws_out, detalhe
        )
        total_linhas     += lin
        total_ignoradas  += ign
        total_retornados += ret
        todos_nao_encontrados |= nao
        wb.close()

    # Formatação
    for col_letter, width in COL_WIDTHS.items():
        ws_out.column_dimensions[col_letter].width = width

    font_padrao = Font(name="Aptos Narrow", size=11)
    for row in ws_out.iter_rows():
        for cell in row:
            cell.font      = font_padrao
            cell.alignment = Alignment(vertical="center", wrap_text=False)

    for col_letter in ["J", "K", "N", "O", "P"]:
        for cell in ws_out[col_letter][1:]:
            if cell.value:
                cell.number_format = "DD/MM/YYYY"

    last_row  = ws_out.max_row
    last_col  = get_column_letter(ws_out.max_column)
    tbl       = XlTable(displayName="Tabela2", ref=f"A1:{last_col}{last_row}")
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,  showColumnStripes=False,
    )
    ws_out.add_table(tbl)
    wb_out.save(OUTPUT_BASE)

    # Grava o horário desta atualização para o front-end exibir "base
    # atualizada em ..." (a data de modificação do arquivo não é confiável
    # quando servida via GitHub Pages/CDN).
    info_path = OUTPUT_BASE.parent / "base_info.json"
    info_path.write_text(
        json.dumps({"atualizado_em": datetime.now().strftime("%d/%m/%Y %H:%M")}, ensure_ascii=False),
        encoding="utf-8",
    )

    # Detalhe mensal por departamento/categoria (com o chamado a chamado)
    # para a tela de Relatórios do portal. Recalculado do zero a cada
    # rodada, já que a extração sempre cobre DATA_INICIO→hoje.
    rel_dir = _DATA_DIR / "relatorios"
    rel_dir.mkdir(exist_ok=True)
    detalhe_json = {
        "atualizado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "departamentos": [
            {
                "nome": depto,
                "meses": [
                    {
                        "mes":      mes,
                        "abertos":  sum(
                            g['atendidosDentro'] + g['atendidosFora'] + g['pendentes']
                            for g in categorias.values()
                        ),
                        "atendidos": sum(
                            g['atendidosDentro'] + g['atendidosFora']
                            for g in categorias.values()
                        ),
                        "pendentes": sum(g['pendentes'] for g in categorias.values()),
                        "categorias": [
                            {
                                "nome":            categoria,
                                "atendidosDentro": g['atendidosDentro'],
                                "atendidosFora":   g['atendidosFora'],
                                "pendentes":       g['pendentes'],
                                "chamados":        g['chamados'],
                            }
                            for categoria, g in sorted(categorias.items())
                        ],
                    }
                    for mes, categorias in sorted(meses.items())
                ],
            }
            for depto, meses in sorted(detalhe.items())
        ],
    }
    (rel_dir / "detalhe_mensal.json").write_text(
        json.dumps(detalhe_json, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    historico_antigo = rel_dir / "historico_mensal.json"
    if historico_antigo.exists():
        historico_antigo.unlink()

    console.print()
    console.print(Panel(
        f"[green]✔[/green] [bold white]{OUTPUT_BASE.name}[/bold white] — "
        f"[bold]{total_linhas}[/bold] registros  |  "
        f"[yellow]{total_ignoradas}[/yellow] ignorados  |  "
        f"[cyan]{total_retornados}[/cyan] retornados",
        title="[green]base.xlsx atualizado[/green]",
        border_style="green",
        padding=(0, 2),
    ))

    if todos_nao_encontrados:
        console.print()
        console.print(
            f"  [yellow]Usuários não encontrados ({len(todos_nao_encontrados)}):[/yellow]"
        )
        for u in sorted(todos_nao_encontrados):
            console.print(f"    [dim]- {u}[/dim]")

# ============================================================
# PUBLICAÇÃO NO GITHUB
# ============================================================

def publicar_no_github():
    console.print()
    console.rule("[bold cyan]Publicando no GitHub[/bold cyan]")
    console.print()

    try:
        with Progress(
            SpinnerColumn(style="cyan"),
            TextColumn("[progress.description]{task.description}"),
            console=console,
            transient=True,
        ) as pg:
            t2 = pg.add_task("[cyan]Sincronizando com GitHub...", total=None)
            pg.update(t2, description="[cyan]Baixando atualizações remotas (pull)...")
            subprocess.run(["git", "-C", REPO_DIR, "pull", "--rebase", "--autostash"],
                           check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            pg.update(t2, description="[cyan]Adicionando arquivos...")
            subprocess.run(["git", "-C", REPO_DIR, "add", "data/"], check=True,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

            diff = subprocess.run(["git", "-C", REPO_DIR, "diff", "--cached", "--quiet"])
            if diff.returncode == 0:
                pg.stop()
                console.print(Panel(
                    "[dim]Nenhuma alteração em data/ desde o último envio.[/dim]",
                    title="[cyan]Nada a publicar[/cyan]",
                    border_style="cyan",
                ))
                return

            pg.update(t2, description="[cyan]Criando commit...")
            msg = f"data: atualização base chamados {agora_str()}"
            subprocess.run(["git", "-C", REPO_DIR, "commit", "-m", msg], check=True,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            pg.update(t2, description="[cyan]Fazendo push...")
            subprocess.run(["git", "-C", REPO_DIR, "push"], check=True,
                           stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

        console.print(Panel(
            "[green]✔[/green] Base enviada para o GitHub!\n"
            "[dim]O Portal de Chamados será atualizado em ~1 minuto.[/dim]\n"
            "[cyan]https://gamaral99.github.io/Chamados/[/cyan]",
            title="[green]GitHub Pages[/green]",
            border_style="green",
            padding=(0, 2),
        ))
    except subprocess.CalledProcessError as e:
        console.print(Panel(
            f"[yellow]base.xlsx atualizado, mas o push falhou:[/yellow]\n[red]{e}[/red]\n\n"
            "[dim]Faça o push manualmente via terminal na pasta do repo.[/dim]",
            title="[yellow]Aviso[/yellow]",
            border_style="yellow",
        ))

# ============================================================
# EXECUÇÃO PRINCIPAL
# ============================================================

def executar(auto=False):
    console.print()
    console.print(Panel(
        Align.center(
            Text.from_markup(
                "[bold cyan]MG CONTÉCNICA[/bold cyan]\n"
                "[dim]Extrator + Atualizador de Chamados[/dim]\n"
                "[bold white]v1.2[/bold white]"
            )
        ),
        border_style="cyan",
        padding=(1, 6),
    ))

    if auto:
        opcao, label_extracao, departamentos = "5", "Todas", None
        headless = True
        console.print(
            "  [dim][--auto] Extração: [bold magenta]Todas (sequencial)[/bold magenta] | "
            "Modo: [bold cyan]Sem janela[/bold cyan][/dim]"
        )
    else:
        opcao, label_extracao, departamentos = perguntar_extracao()
        headless = perguntar_modo()

    data_inicio = DATA_INICIO
    data_fim    = hoje_str()

    if opcao == "5":
        extracoes = [(v["label"], v["departamentos"]) for v in TIPOS_EXTRACAO.values()]
    else:
        extracoes = [(label_extracao, departamentos)]

    console.print()
    if opcao == "5":
        console.print(
            f"  [dim]Extração: [bold cyan]Todas (sequencial)[/bold cyan] | "
            f"[bold cyan]{len(extracoes)}[/bold cyan] extrações[/dim]"
        )
    else:
        console.print(
            f"  [dim]Extração: [bold cyan]{label_extracao}[/bold cyan] | "
            f"Departamentos: [bold cyan]{len(departamentos)}[/bold cyan][/dim]"
        )
    console.print(
        f"  [dim]Período: [bold cyan]{data_inicio}[/bold cyan] → "
        f"[bold cyan]{data_fim}[/bold cyan][/dim]"
    )
    console.print()
    console.rule("[bold cyan]Executando[/bold cyan]")
    console.print()

    driver           = None
    arquivos_gerados = []

    with Progress(
        SpinnerColumn(style="cyan"),
        TextColumn("[progress.description]{task.description}"),
        TimeElapsedColumn(),
        console=console,
        transient=True,
    ) as progress:
        task = progress.add_task("[cyan]Iniciando...", total=None)

        try:
            progress.update(task, description="[cyan]Iniciando Chrome...")
            driver = iniciar_driver(PASTA_DOWNLOAD, headless=headless)

            limpar_pasta_download(PASTA_DOWNLOAD, progress, task)

            fazer_login(driver, progress, task)
            navegar_para_relatorio(driver, progress, task)

            deptos_anteriores = 0
            for i, (lbl, deptos) in enumerate(extracoes):
                if i > 0:
                    limpar_departamentos(driver, deptos_anteriores, progress, task)

                progress.update(task,
                                description=f"[cyan]Extração {i+1}/{len(extracoes)}: {lbl}...")

                selecionar_departamentos(driver, deptos, progress, task)
                preencher_datas(driver, data_inicio, data_fim, progress, task)
                marcar_encerrados(driver, progress, task)

                arquivo_bruto = gerar_relatorio(driver, PASTA_DOWNLOAD, progress, task)

                if arquivo_bruto:
                    progress.update(task, description="[cyan]Renomeando arquivo...")
                    _, novo_nome = renomear_arquivo(arquivo_bruto, PASTA_DOWNLOAD, label=lbl)
                    arquivos_gerados.append((lbl, novo_nome))
                else:
                    arquivos_gerados.append((lbl, None))

                deptos_anteriores = len(deptos)

        except Exception as erro:
            progress.stop()
            console.print()
            console.print(Panel(
                f"[red]{erro}[/red]",
                title="[red]Erro durante a execução[/red]",
                border_style="red",
            ))
            console.print_exception()

        finally:
            if driver:
                driver.quit()

    # ── Resultado dos downloads ───────────────────────────────
    console.print()

    sucesso = [a for a in arquivos_gerados if a[1] is not None]
    falha   = [a for a in arquivos_gerados if a[1] is None]

    if sucesso:
        console.rule("[bold green]Downloads concluídos[/bold green]")
        console.print()

        info = Table(box=box.ROUNDED, border_style="dim", show_header=False, padding=(0, 2))
        info.add_column("Label", style="dim",       width=20)
        info.add_column("Valor", style="bold white")
        info.add_row("Data início", data_inicio)
        info.add_row("Data fim",    data_fim)
        info.add_row("Pasta",       PASTA_DOWNLOAD)
        console.print(Align.center(info))
        console.print()

        for lbl, nome in sucesso:
            console.print(Panel(
                f"[green]✔[/green] [bold white]{nome}[/bold white]",
                title=f"[green]{lbl}[/green]",
                border_style="green",
                padding=(0, 2),
            ))

    if falha:
        console.print()
        for lbl, _ in falha:
            console.print(Panel(
                "[yellow]Download não detectado dentro do tempo limite.[/yellow]",
                title=f"[yellow]Aviso — {lbl}[/yellow]",
                border_style="yellow",
            ))

    if not arquivos_gerados:
        console.rule("[bold yellow]Concluído com aviso[/bold yellow]")
        console.print()
        console.print(Panel(
            "[yellow]Nenhum download detectado.[/yellow]",
            title="[yellow]Aviso[/yellow]",
            border_style="yellow",
        ))

    # ── Atualizar base.xlsx ───────────────────────────────────
    if sucesso:
        try:
            atualizar_base()
            publicar_no_github()
        except Exception as erro:
            console.print()
            console.print(Panel(
                f"[red]{erro}[/red]",
                title="[red]Erro ao atualizar base.xlsx[/red]",
                border_style="red",
            ))

    console.print()

# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extrator + Atualizador de Base de Chamados MG Contécnica"
    )
    parser.add_argument(
        "--auto",
        action="store_true",
        help="Execução automática: extrai Todas (sequencial), modo headless, "
             "sem prompts, e publica no GitHub ao final.",
    )
    args = parser.parse_args()
    executar(auto=args.auto)
