"""
atualizar_base.py
─────────────────
Lê os relatórios de chamados (Geral, GC, Controladoria e Paralegal) da
pasta Att Base, substitui usuários AD das colunas de responsável/
solicitante pelo nome de exibição, resolve o Coordenador de cada
responsável (via Coordenadores.xlsx) e filtra status 'Entregue Ao
Solicitante'/'Encerrado' — EXCETO para registros cujo Departamento
Solicitante seja de Gerência de Contas (SP/Santos/RJ), que são salvos
com a flag Retornado = SIM na coluna Retornado.

Mesmo schema de colunas e mesmas regras de negócio do pipeline real
(backend/extrato_chamados.py, rodado pela tarefa agendada) — este
script processa arquivos já baixados manualmente, sem depender do
Selenium/login. Não gera o detalhe mensal (data/relatorios/
detalhe_mensal.json) usado na tela de Relatórios — isso continua
exclusivo do extrato_chamados.py.

Estrutura esperada:
  Portal-dos-chamados/
  ├── data/
  │   ├── base.xlsx                          ← será sobrescrito
  │   ├── base_info.json                     ← será sobrescrito
  │   └── Att Base/
  │       ├── Chamados - Geral - *.xlsx
  │       ├── Chamados - GC - *.xlsx
  │       ├── Chamados - Controladoria - *.xlsx
  │       ├── Chamados - Paralegal - *.xlsx  (opcional)
  │       ├── Relatório Colaboradores Ativos_*.xlsx
  │       └── Coordenadores*.xlsx            (opcional)
  └── backend/atualizar_base.py
"""

import os
import re
import sys
import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ─── CAMINHOS ────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR / "data" if (BASE_DIR / "data").exists() else BASE_DIR.parent / "data"
ATT_DIR  = DATA_DIR / "Att Base"
OUTPUT   = DATA_DIR / "base.xlsx"

# ─── DEPARTAMENTOS QUE SALVAM RETORNADOS ─────────────────────────────
# Inclui os grupos de GC de Santos e Rio de Janeiro — sem eles, um chamado
# "Entregue Ao Solicitante" vindo desses departamentos era tratado como
# comum e descartado (em vez de mantido com Retornado = SIM), fazendo os
# relatórios de Santos/RJ parecerem sem atualização.
DEPTS_RETORNADOS = {
    "gerencia de contas", "gc - administrativo",
    "santos - gc", "santos - adm",
    "rj - gc", "rj - gc administrativo",
}
STATUS_ENTREGUE  = "entregue ao solicitante"
STATUS_ENCERRADO = "encerrado"
STATUS_DEVOLVIDO_SOLICITANTE = "devolvido para solicitante"

# ─── MAPEAMENTO DE COLUNAS (índice 0-based) ──────────────────────────
COL_MAP = [
    (0,  0),   # A → A  | Id
    (1,  1),   # B → B  | Categoria
    (2,  2),   # C → C  | Assunto
    (3,  3),   # D → D  | Departamento Solicitante
    (4,  4),   # E → E  | Solicitação
    (5,  5),   # F → F  | IdCliente
    (6,  6),   # G → G  | Cliente
    (8,  7),   # I → H  | Responsável (com lookup)
    (9,  8),   # J → I  | Departamento Responsavel
    (13, 9),   # N → J  | Data Cadastro
    (14, 10),  # O → K  | Prazo Vencimento
    (15, 11),  # P → L  | Status
    (16, 12),  # Q → M  | Solicitante (com lookup)
    (17, 13),  # R → N  | Inicio Atend.
    (18, 14),  # S → O  | DataPrevisaoAtendimento
    (19, 15),  # T → P  | Data Entrega
    (20, 16),  # U → Q  | Responsável pela conclusão
    (21, 17),  # V → R  | Ultimo Comentário
]

LOOKUP_COLS = {8, 16}  # I e Q do relatório de origem

HEADERS = [
    "Id", "Categoria", "Assunto", "Departamento Solicitante",
    "Solicitação", "IdCliente", "Cliente", "Responsável",
    "Departamento Responsavel", "Data Cadastro", "Prazo Vencimento",
    "Status", "Solicitante", "Inicio Atend.", "DataPrevisaoAtendimento",
    "Data Entrega", "Responsável pela conclusão", "Ultimo Comentário",
    "Retornado", "Departamento Responsavel Original", "Coordenador",
]

COL_WIDTHS = {
    'A': 10,    'B': 11.86, 'C': 10.43, 'D': 26.29,
    'E': 13.0,  'F': 11.57, 'G': 9.86,  'H': 14.71,
    'I': 28.14, 'J': 15.86, 'K': 19.29, 'L': 12.86,
    'M': 12.86, 'N': 14.43, 'O': 14.43, 'P': 15.57,
    'Q': 28.57, 'R': 20.29, 'S': 12.0,  'T': 28.14,
    'U': 22.0,
}


# ─── HELPERS ─────────────────────────────────────────────────────────
def find_file(folder, pattern, required=True):
    matches = list(folder.glob(pattern))
    if not matches:
        if required:
            raise FileNotFoundError(f"Nenhum arquivo encontrado com padrão '{pattern}' em {folder}")
        return None
    return max(matches, key=os.path.getmtime)

def find_all(folder, pattern):
    return list(folder.glob(pattern))

def norm_nome(nome):
    if not nome:
        return ""
    return re.sub(r"\s+", " ", str(nome).strip().upper())


def processar_relatorio(ws, label, linhas_total, ignoradas_total,
                        retornados_total, nao_encontrados, ws_out,
                        colab_map, coordenador_map):
    linhas     = 0
    ignoradas  = 0
    retornados = 0

    def resolver_nome(valor):
        if not valor:
            return valor
        chave = str(valor).strip().lower()
        return colab_map.get(chave, str(valor).strip())

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue

        status_val = str(row[15]).strip().lower() if row[15] else ''
        dept_sol   = str(row[3]).strip().lower()  if row[3]  else ''

        # "Encerrado" conta como baixado igual a "Entregue Ao Solicitante",
        # mas NUNCA vira "Retornado" — só chamados entregues normalmente
        # entram como retornado. Encerrado é sempre descartado da base.
        eh_entregue = (status_val == STATUS_ENTREGUE)
        eh_encerrado = (status_val == STATUS_ENCERRADO)
        eh_fechado  = eh_entregue or eh_encerrado
        eh_gc       = (dept_sol in DEPTS_RETORNADOS)
        manter_como_retornado = eh_entregue and eh_gc

        if eh_fechado and not manter_como_retornado:
            ignoradas += 1
            continue

        nova_linha = [None] * len(HEADERS)

        for col_src, col_dst in COL_MAP:
            valor = row[col_src] if col_src < len(row) else None
            if col_src in LOOKUP_COLS:
                nome_resolvido = resolver_nome(valor)
                bruto = str(valor).strip() if valor else ''
                if valor and nome_resolvido == bruto and ' ' not in bruto:
                    nao_encontrados.add(bruto)
                valor = nome_resolvido
            nova_linha[col_dst] = valor

        # "Devolvido para Solicitante": a pendência passa a ser do depto
        # solicitante (coluna D), não mais do depto responsável (coluna I).
        # O depto responsável original é preservado na coluna de rastreabilidade.
        if status_val == STATUS_DEVOLVIDO_SOLICITANTE:
            nova_linha[19] = nova_linha[8]
            nova_linha[8]  = nova_linha[3]

        nova_linha[18] = "SIM" if manter_como_retornado else "NÃO"
        if manter_como_retornado:
            retornados += 1

        # Coordenador do responsável pelo chamado (via Coordenadores.xlsx,
        # cruzado pelo Nome de Exibição já resolvido na coluna H).
        nova_linha[20] = coordenador_map.get(norm_nome(nova_linha[7]), "")

        ws_out.append(nova_linha)
        linhas += 1

    print(f"  {label}: {linhas} incluídos  |  {ignoradas} ignorados  |  retornados: {retornados}")
    return linhas_total + linhas, ignoradas_total + ignoradas, retornados_total + retornados


# ─── MAIN ────────────────────────────────────────────────────────────
def main():
    print("Localizando arquivos...")
    colab_path = find_file(ATT_DIR, "*olaboradores*.xlsx")
    coord_path = find_file(ATT_DIR, "*oordenadores*.xlsx",            required=False)

    # Arquivos específicos: GC, Controladoria e Paralegal (pega o mais recente de cada)
    gc_path  = find_file(ATT_DIR, "*Chamados - GC*.xlsx",             required=False)
    ctr_path = find_file(ATT_DIR, "*Chamados - Controladoria*.xlsx",  required=False)
    pl_path  = find_file(ATT_DIR, "*Chamados - Paralegal*.xlsx",      required=False)

    gc_name  = gc_path.name  if gc_path  else None
    ctr_name = ctr_path.name if ctr_path else None
    pl_name  = pl_path.name  if pl_path  else None

    # Chamados gerais: todos com "Chamados" no nome, exceto GC/Controladoria/Paralegal
    chamados_paths = [
        p for p in find_all(ATT_DIR, "*Chamados*.xlsx")
        if p.name != gc_name and p.name != ctr_name and p.name != pl_name
    ]

    if not chamados_paths:
        raise FileNotFoundError(f"Nenhum arquivo *Chamados*.xlsx (Geral) encontrado em {ATT_DIR}")

    print(f"  Colaboradores: {colab_path.name}")
    print(f"  Coordenadores: {coord_path.name if coord_path else 'nao encontrado'}")
    for p in chamados_paths:
        print(f"  Chamados:      {p.name}")
    print(f"  GC:            {gc_name  or 'nao encontrado'}")
    print(f"  CTR:           {ctr_name or 'nao encontrado'}")
    print(f"  Paralegal:     {pl_name  or 'nao encontrado'}")

    # Carregar mapa de colaboradores
    print("\nCarregando colaboradores...")
    wb_colab  = openpyxl.load_workbook(colab_path, read_only=True, data_only=True)
    ws_colab  = wb_colab.active
    colab_map = {}
    for row in ws_colab.iter_rows(min_row=2, values_only=True):
        nome_exib = row[1]
        usuario   = row[2]
        if usuario and nome_exib:
            colab_map[str(usuario).strip().lower()] = str(nome_exib).strip()
    wb_colab.close()
    print(f"  {len(colab_map)} colaboradores carregados.")

    # Carregar mapa de coordenadores (Nome de Exibição → Coordenador).
    # Opcional — se o arquivo não existir, a coluna Coordenador fica em branco.
    coordenador_map = {}
    if coord_path:
        wb_coord = openpyxl.load_workbook(coord_path, read_only=True, data_only=True)
        for row in wb_coord.active.iter_rows(min_row=2, values_only=True):
            nome_exib   = row[1] if len(row) > 1 else None
            coordenador = row[3] if len(row) > 3 else None
            if nome_exib and coordenador:
                coordenador_map[norm_nome(nome_exib)] = str(coordenador).strip()
        wb_coord.close()
    print(f"  {len(coordenador_map)} coordenadores carregados.")

    # Preparar workbook de saída
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Base"
    ws_out.append(HEADERS)

    linhas_total     = 0
    ignoradas_total  = 0
    retornados_total = 0
    nao_encontrados  = set()

    print("\nProcessando relatórios...")
    fontes = [(p.stem, p) for p in sorted(chamados_paths)]
    if gc_path:  fontes.append(("GC",            gc_path))
    if ctr_path: fontes.append(("Controladoria", ctr_path))
    if pl_path:  fontes.append(("Paralegal",     pl_path))

    for label, path in fontes:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        linhas_total, ignoradas_total, retornados_total = processar_relatorio(
            ws, label, linhas_total, ignoradas_total, retornados_total,
            nao_encontrados, ws_out, colab_map, coordenador_map
        )
        wb.close()

    # Formatação
    for col_letter, width in COL_WIDTHS.items():
        ws_out.column_dimensions[col_letter].width = width

    font_padrao = Font(name='Aptos Narrow', size=11)
    for row in ws_out.iter_rows():
        for cell in row:
            cell.font      = font_padrao
            cell.alignment = Alignment(vertical='center', wrap_text=False)

    for col_letter in ['J', 'K', 'N', 'O', 'P']:
        for cell in ws_out[col_letter][1:]:
            if cell.value:
                cell.number_format = 'DD/MM/YYYY'

    last_row  = ws_out.max_row
    last_col  = get_column_letter(ws_out.max_column)
    table_ref = f"A1:{last_col}{last_row}"
    table     = Table(displayName="Tabela2", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False, showLastColumn=False,
        showRowStripes=True,  showColumnStripes=False,
    )
    ws_out.add_table(table)

    wb_out.save(OUTPUT)

    # Grava o horário desta atualização para o front-end exibir "base
    # atualizada em ..." (a data de modificação do arquivo não é confiável
    # quando servida via GitHub Pages/CDN) — mesmo arquivo que o
    # extrato_chamados.py mantém.
    info_path = OUTPUT.parent / "base_info.json"
    info_path.write_text(
        json.dumps({"atualizado_em": datetime.now().strftime("%d/%m/%Y %H:%M")}, ensure_ascii=False),
        encoding="utf-8",
    )

    print(f"\nbase.xlsx atualizado com {linhas_total} registros -> {OUTPUT}")
    print(f"{ignoradas_total} registro(s) ignorado(s).")
    print(f"{retornados_total} registro(s) marcado(s) como Retornado.")

    if nao_encontrados:
        print(f"\nUsuarios nao encontrados no relatorio de colaboradores ({len(nao_encontrados)}):")
        for u in sorted(nao_encontrados):
            print(f"   - {u}")
    else:
        print("Todos os usuarios foram resolvidos com sucesso.")


if __name__ == "__main__":
    main()
